# app/excel/excel_io.py
"""
Excel import/export helpers.
"""

from __future__ import annotations

import unicodedata
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMNS = [
    "ID SUBASTA",
    "ITEM",
    "DESCRIPCION",
    "UNIDAD DE MEDIDA",
    "CANTIDAD",
    "MARCA",
    "OBS USUARIO",
    "CONVERSIÓN USD",
    "COSTO UNIT USD",
    "COSTO TOTAL USD",
    "COSTO UNIT ARS",
    "COSTO TOTAL ARS",
    "RENTA MINIMA %",
    "PRECIO UNIT ACEPTABLE",
    "PRECIO TOTAL ACEPTABLE",
    "PRECIO DE REFERENCIA",
    "PRECIO REF UNITARIO",
    "RENTA REFERENCIA %",
    "MEJOR OFERTA ACTUAL",
    "OFERTA PARA MEJORAR",
    "PRECIO UNIT MEJORA",
    "RENTA PARA MEJORAR %",
    "OBS / CAMBIO",
]

USER_FIELDS = {
    "UNIDAD DE MEDIDA",
    "MARCA",
    "OBS USUARIO",
    "CONVERSIÓN USD",
    "COSTO UNIT ARS",
    "COSTO TOTAL ARS",
    "RENTA MINIMA %",
}

CALC_FIELDS = {
    "COSTO UNIT USD",
    "COSTO TOTAL USD",
    "PRECIO UNIT ACEPTABLE",
    "PRECIO TOTAL ACEPTABLE",
    "PRECIO REF UNITARIO",
    "RENTA REFERENCIA %",
    "PRECIO UNIT MEJORA",
    "RENTA PARA MEJORAR %",
}

PLAYWRIGHT_FIELDS = {
    "ID SUBASTA",
    "ITEM",
    "DESCRIPCION",
    "CANTIDAD",
    "PRECIO DE REFERENCIA",
    "MEJOR OFERTA ACTUAL",
    "OFERTA PARA MEJORAR",
    "OBS / CAMBIO",
}

OBS_FIELDS = {
    "ID SUBASTA",
    "ITEM",
    "DESCRIPCION",
    "SUBTOTAL PARA MEJORAR",
    "Precio referencia",
}

TABLE_NAME = "T_Subastas"
SHEET_NAME = "Subastas"

MONEY_COLS = {
    "COSTO UNIT USD",
    "COSTO TOTAL USD",
    "COSTO UNIT ARS",
    "COSTO TOTAL ARS",
    "PRECIO UNIT ACEPTABLE",
    "PRECIO TOTAL ACEPTABLE",
    "PRECIO DE REFERENCIA",
    "PRECIO REF UNITARIO",
    "PRECIO UNIT MEJORA",
    "OFERTA PARA MEJORAR",
}

PERCENT_COLS = {
    "RENTA MINIMA %",
    "RENTA REFERENCIA %",
    "RENTA PARA MEJORAR %",
}

# Se mantiene por referencia conceptual, no se usa para escribir fórmulas
FORMULAS = {
    "COSTO UNIT USD": "=[@[COSTO UNIT ARS]]/[@[CONVERSIÓN USD]]",
    "COSTO TOTAL USD": "=[@[COSTO TOTAL ARS]]/[@[CONVERSIÓN USD]]",
    "PRECIO UNIT ACEPTABLE": "=(1+[@[RENTA MINIMA %]])*[@[COSTO UNIT ARS]]",
    "PRECIO TOTAL ACEPTABLE": "=(1+[@[RENTA MINIMA %]])*[@[COSTO TOTAL ARS]]",
    "PRECIO REF UNITARIO": "=[@[PRECIO DE REFERENCIA]]/[@[CANTIDAD]]",
    "RENTA REFERENCIA %": "=([@[PRECIO DE REFERENCIA]]/[@[COSTO TOTAL ARS]])-1",
    "PRECIO UNIT MEJORA": "=[@[OFERTA PARA MEJORAR]]/[@[CANTIDAD]]",
    "RENTA PARA MEJORAR %": "=([@[OFERTA PARA MEJORAR]]/[@[COSTO TOTAL ARS]])-1",
}


def _normalize_header(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\u00a0", " ")
    text = " ".join(text.split())
    return text.upper()


def export_subasta_to_excel(*, rows: Iterable[dict], out_path: str) -> None:
    """
    Exporta filas a Excel creando una tabla con formato y fórmulas.
    El archivo queda completamente editable (sin protección).
    """
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header
    ws.append(COLUMNS)

    # Data
    if rows:
        for row in rows:
            ws.append([row.get(col) for col in COLUMNS])
    else:
        # Fila vacía para que existan tabla y fórmulas
        ws.append([None for _ in COLUMNS])

    max_row = ws.max_row
    max_col = ws.max_column

    # Tabla
    ref = f"{ws.cell(row=1, column=1).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
    table = Table(displayName=TABLE_NAME, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Índices y letras
    header_index = {name: i + 1 for i, name in enumerate(COLUMNS)}
    header_letter = {name: get_column_letter(i) for name, i in header_index.items()}

    # Formatos
    for col_name, col_idx in header_index.items():
        if col_name in MONEY_COLS:
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "$ #.##0,00"
        if col_name == "RENTA MINIMA %":
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "0,00"
        elif col_name in PERCENT_COLS:
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "0,00%"

    wb.save(out_path)


def import_excel_to_rows(*, file_path: str) -> list[dict]:
    wb = load_workbook(file_path, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    headers = [c.value for c in ws[1]]
    header_map = {_normalize_header(h): h for h in headers if h}
    
    # Requeridas (sin acentos, consistente con _normalize_header)
    required = {
        "ID SUBASTA",
        "ITEM",
        "UNIDAD DE MEDIDA",
        "MARCA",
        "OBS USUARIO",
        "CONVERSION USD",
        "COSTO UNIT ARS",
        "COSTO TOTAL ARS",
        "RENTA MINIMA %",
    }
    if not required.issubset(set(header_map.keys())):
        missing = sorted(required - set(header_map.keys()))
        raise ValueError(f"Faltan columnas requeridas: {missing}")

    header_index = {
        _normalize_header(name): i for i, name in enumerate(headers) if name
    }
    reverse_headers = {
        _normalize_header(name): name for name in COLUMNS
    }

    rows: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        
        # Construir fila solo con USER_FIELDS (ignorar valores calculados y de Playwright)
        row_raw = {
            reverse_headers.get(norm, norm): r[idx] if idx < len(r) else None
            for norm, idx in header_index.items()
        }
        
        # Filtrar: solo incluir ID, ITEM (para identificación) + USER_FIELDS
        row = {}
        for name in ["ID SUBASTA", "ITEM"]:
            if name in row_raw:
                row[name] = row_raw[name]
        
        for name in USER_FIELDS:
            if name in row_raw:
                row[name] = row_raw[name]
        
        rows.append(row)

    return rows
