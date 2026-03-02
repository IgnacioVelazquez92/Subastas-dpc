#!/usr/bin/env python3
"""
Test que verifica la coherencia de import_excel_to_rows:
- Crea un Excel con todas las columnas
- Importa y verifica que solo contiene USER_FIELDS (+ ID/ITEM)
"""
from openpyxl import Workbook
from app.excel.excel_io import (
    COLUMNS, USER_FIELDS, CALC_FIELDS, PLAYWRIGHT_FIELDS,
    import_excel_to_rows
)

def test_import_filters():
    # Crear workbook con todas las columnas
    wb = Workbook()
    ws = wb.active
    ws.title = "Subastas"
    
    # Headers
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Data row 1: valores de test
    test_data = {
        "ID SUBASTA": "123456",
        "ITEM": "1",
        "DESCRIPCION": "Test Item",
        "UNIDAD DE MEDIDA": "KG",
        "CANTIDAD": 100,
        "ITEMS POR RENGLON": 5,
        "MARCA": "TestBrand",
        "OBS USUARIO": "Test obs",
        "CONVERSIÓN USD": 1500,
        "COSTO UNIT USD": 5000,       # CALC_FIELD - no debería importar
        "COSTO TOTAL USD": 500000,    # CALC_FIELD - no debería importar
        "COSTO UNIT ARS": 7500000,
        "COSTO TOTAL ARS": 750000000,
        "RENTA MINIMA %": 1.5,
        "PRECIO UNIT ACEPTABLE": 11250000,  # CALC_FIELD
        "PRECIO TOTAL ACEPTABLE": 1125000000,  # CALC_FIELD
        "PRESUPUESTO OFICIAL": 1000000000,  # PLAYWRIGHT_FIELD
        "PRECIO DE REFERENCIA": 50000000,  # CALC_FIELD
        "RENTA REFERENCIA %": 0.33,  # CALC_FIELD
        "MEJOR OFERTA ACTUAL": 950000000,  # PLAYWRIGHT_FIELD
        "OFERTA PARA MEJORAR": 900000000,  # PLAYWRIGHT_FIELD
        "PRECIO UNIT MEJORA": 9000000,  # CALC_FIELD
        "RENTA PARA MEJORAR %": 0.2,  # CALC_FIELD
        "OBS / CAMBIO": "Change log",  # PLAYWRIGHT_FIELD
    }
    
    for col_idx, col_name in enumerate(COLUMNS, 1):
        value = test_data.get(col_name, None)
        ws.cell(row=2, column=col_idx, value=value)
    
    # Guardar
    import tempfile
    import os
    test_file = os.path.join(tempfile.gettempdir(), "test_import.xlsx")
    wb.save(test_file)
    print(f"✅ Archivo test creado: {test_file}")
    
    # Importar
    try:
        imported_rows = import_excel_to_rows(file_path=test_file)
    except Exception as e:
        print(f"❌ Error importando: {e}")
        raise AssertionError(f"Error importando: {e}") from e
    
    if not imported_rows:
        raise AssertionError("No se importaron filas")
    
    row = imported_rows[0]
    print(f"\n📋 Campos en fila importada ({len(row)} campos):")
    for key in sorted(row.keys()):
        print(f"  ✓ {key}")
    
    # Validación
    print(f"\n🔍 Validación:")
    
    # Debe contener ID/ITEM + USER_FIELDS
    required = {"ID SUBASTA", "ITEM"} | USER_FIELDS
    actual = set(row.keys())
    
    missing = required - actual
    if missing:
        raise AssertionError(f"Faltan campos requeridos: {missing}")
    
    # NO debe contener CALC_FIELDS (excepto ID/ITEM)
    bad_calc = actual & CALC_FIELDS
    if bad_calc:
        raise AssertionError(f"Contiene CALC_FIELDS (no debería): {bad_calc}")
    
    # NO debe contener PLAYWRIGHT_FIELDS (excepto ID/ITEM)
    bad_pw = (actual & PLAYWRIGHT_FIELDS) - {"ID SUBASTA", "ITEM"}
    if bad_pw:
        raise AssertionError(f"Contiene PLAYWRIGHT_FIELDS (no debería): {bad_pw}")
    
    # Validar valores correctos
    print(f"\n📊 Validación de valores:")
    
    tests = [
        ("CONVERSIÓN USD", 1500, "debe ser 1500"),
        ("COSTO UNIT ARS", 7500000, "debe ser 7500000"),
        ("COSTO TOTAL ARS", 750000000, "debe ser 750000000"),
        ("RENTA MINIMA %", 1.5, "debe ser 1.5"),
    ]
    
    for field, expected, note in tests:
        actual_val = row.get(field)
        status = "✓" if actual_val == expected else "✗"
        print(f"  {status} {field}: {actual_val} (esperado: {expected}) - {note}")
    
    # Validar que CALC_FIELDS NO tienen valor
    print(f"\n🚫 CALC_FIELDS no deben estar presentes:")
    for calc_field in CALC_FIELDS:
        if calc_field in row:
            raise AssertionError(f"{calc_field} está presente y no debería")
        print(f"  ✓ {calc_field} ausente")
    
    print(f"\n✅ ÉXITO: Import solo contiene USER_FIELDS + identificadores")

if __name__ == "__main__":
    import sys
    test_import_filters()
    sys.exit(0)
